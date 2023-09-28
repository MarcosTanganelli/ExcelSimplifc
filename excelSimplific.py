import openpyxl

def converter_para_float(valor):
    # Substitui vírgula por ponto e converte para float
    valor = valor.replace('.', '') #5.737,70 -> valores como esse da erro, então substitui para 5737,70
    valor = valor.replace(',', '.')#5737,70 ->5737.70 -> esse dado o python aceita
    return float(valor)

def excel(Composicoes_Analitico, Insumos, save):
    # Abra o arquivo Excel
    workbook = openpyxl.load_workbook(Composicoes_Analitico)
    workbook_novo = openpyxl.Workbook()


    planilha_nova = workbook_novo.create_sheet('DADOS')
    planilha1 = workbook.active 

    # Determine a última linha 
    ultima_linha = planilha1.max_row

    # Loop por todas as 
    for i in range(8, ultima_linha + 1): 
        if not planilha1.cell(i, 13).value:  # caso não tenha subcode é um grupo
            planilha_nova.cell(i-6, 1, "X") #grupo
            planilha_nova.cell(i-6, 2, planilha1.cell(i, 7).value)#cod
            planilha_nova.cell(i-6, 3, planilha1.cell(i, 8).value)#descrição
            planilha_nova.cell(i-6, 4, planilha1.cell(i, 9).value)#unidade
            planilha_nova.cell(i-6, 5, planilha1.cell(i, 17).value)#coeficiente
            planilha_nova.cell(i-6, 6, planilha1.cell(i, 22).value)#material
            valor1 = converter_para_float(planilha1.cell(i, 11).value)  # Converte o valor com vírgula para float
            valor2 = converter_para_float(planilha1.cell(i, 22).value)  # Converte o valor com vírgula para float
            planilha_nova.cell(i-6, 7, valor1 - valor2)#Mão de obra
        else:
            planilha_nova.cell(i-6, 2, planilha1.cell(i, 13).value)#cod
            planilha_nova.cell(i-6, 3, planilha1.cell(i, 14).value)#descrição
            planilha_nova.cell(i-6, 4, planilha1.cell(i, 15).value)#unidade
            planilha_nova.cell(i-6, 5, planilha1.cell(i, 17).value)#coeficiente
            planilha_nova.cell(i-6, 8, planilha1.cell(i, 18).value)#preço unitario
            planilha_nova.cell(i-6, 9, planilha1.cell(i, 19).value)#preço total


    # Atualize as etiquetas de cabeçalho
    planilha_nova.cell(1, 1, "GRUPO")
    planilha_nova.cell(1, 2, "CODIGO")
    planilha_nova.cell(1, 3, "DESCRIÇÃO")
    planilha_nova.cell(1, 4, "UNIDADE")
    planilha_nova.cell(1, 5, "COEFICIENTE")
    planilha_nova.cell(1, 6, "MATERIAL")
    planilha_nova.cell(1, 7, "MÃO DE OBRA")
    planilha_nova.cell(1, 8, "PREÇO UNITÁRIO")
    planilha_nova.cell(1, 9, "PREÇO TOTAL")
    
    workbook1 = openpyxl.load_workbook(Insumos)
    planilha1_segundo = workbook1.active 
    ultima_linha_2 = planilha1_segundo.max_row
    ultima_linha_p2 = planilha_nova.max_row - 7

    for i in range(8, ultima_linha_2 -2): 
        planilha_nova.cell(i+ultima_linha_p2, 2, planilha1_segundo.cell(i, 1).value)#cod
        planilha_nova.cell(i+ultima_linha_p2, 3, planilha1_segundo.cell(i, 2).value)#descrição
        planilha_nova.cell(i+ultima_linha_p2, 4, planilha1_segundo.cell(i, 3).value)#unidade
        planilha_nova.cell(i+ultima_linha_p2, 8, planilha1_segundo.cell(i, 5).value)#preço unitario

    # Salve as alterações
    workbook_novo.save(save) 
    #46601
    return 1
